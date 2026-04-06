"""
AIGI-Holmes backend — export detection reports to PDF / Excel.
"""

import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont


def _register_chinese_font() -> str:
    """Try to register a Chinese-capable font; fall back to Helvetica."""
    import os
    candidates = [
        os.path.join(os.environ.get("WINDIR", r"C:\Windows"), "Fonts", "msyh.ttc"),
        os.path.join(os.environ.get("WINDIR", r"C:\Windows"), "Fonts", "simsun.ttc"),
        "/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc",
        "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
    ]
    for path in candidates:
        if os.path.isfile(path):
            try:
                pdfmetrics.registerFont(TTFont("ChineseFont", path))
                return "ChineseFont"
            except Exception:
                continue
    return "Helvetica"


def _draw_page_decoration(canvas_obj, doc):
    """Diagonal watermark text + footer on every page."""
    page_w, page_h = A4

    # Diagonal watermark
    canvas_obj.saveState()
    canvas_obj.setFont("Helvetica", 54)
    canvas_obj.setFillColorRGB(0.91, 0.91, 0.91)
    canvas_obj.translate(page_w / 2, page_h / 2)
    canvas_obj.rotate(40)
    canvas_obj.drawCentredString(0, 0, "AIGI-Holmes")
    canvas_obj.restoreState()

    # Footer
    canvas_obj.saveState()
    canvas_obj.setFont("Helvetica", 7.5)
    canvas_obj.setFillColorRGB(0.55, 0.55, 0.55)
    pn = canvas_obj.getPageNumber()
    canvas_obj.drawString(
        20 * mm, 10 * mm,
        f"AIGI-Holmes \u81ea\u52a8\u68c0\u6d4b\u62a5\u544a  \u00b7  \u7b2c {pn} \u9875  \u00b7  \u672c\u62a5\u544a\u4ec5\u4f9b\u53c2\u8003\uff0c\u4e0d\u4f5c\u4e3a\u6700\u7ec8\u5224\u5b9a\u4f9d\u636e"
    )
    canvas_obj.restoreState()


def export_pdf(report: dict) -> bytes:
    """Generate a PDF report with watermark, verdict banner and analysis section."""
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        topMargin=22 * mm, bottomMargin=24 * mm,
        leftMargin=20 * mm, rightMargin=20 * mm,
    )

    font_name = _register_chinese_font()
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(
        name="CN", fontName=font_name, fontSize=10, leading=15,
    ))
    styles.add(ParagraphStyle(
        name="CNSmall", fontName=font_name, fontSize=9, leading=14,
        textColor=colors.Color(0.45, 0.45, 0.45),
    ))
    styles.add(ParagraphStyle(
        name="CNTitle", fontName=font_name, fontSize=17, leading=24,
        spaceAfter=4, textColor=colors.black,
    ))
    styles.add(ParagraphStyle(
        name="CNSection", fontName=font_name, fontSize=12, leading=18,
        spaceBefore=10, spaceAfter=5, textColor=colors.black,
    ))

    is_fake = (report.get("label", "") == "FAKE")
    col_w = [100, 360]   # label column / value column (mm → pt auto-scaled)

    elements: list = []

    # ── Title ────────────────────────────────────────────────────
    elements.append(Paragraph("AIGI-Holmes \u56fe\u50cf\u68c0\u6d4b\u62a5\u544a", styles["CNTitle"]))
    elements.append(Spacer(1, 3 * mm))

    # ── Verdict banner ───────────────────────────────────────────
    verdict_bg  = colors.Color(0.96, 0.91, 0.91) if is_fake else colors.Color(0.90, 0.96, 0.90)
    verdict_fg  = colors.Color(0.78, 0.10, 0.10) if is_fake else colors.Color(0.04, 0.54, 0.18)
    verdict_txt = "\u26a0 AI\u751f\u6210\u56fe\u50cf" if is_fake else "\u2713 \u771f\u5b9e\u7167\u7247"
    conf_val    = report.get("confidence", 0)

    banner_row = [[
        Paragraph(
            f'<font color="#{int(verdict_fg.red*255):02x}'
            f'{int(verdict_fg.green*255):02x}'
            f'{int(verdict_fg.blue*255):02x}"><b>{verdict_txt}</b></font>',
            styles["CN"],
        ),
        Paragraph(f'\u7f6e\u4fe1\u5ea6\uff1a<b>{conf_val:.1f}%</b>', styles["CN"]),
    ]]
    banner = Table(banner_row, colWidths=[320, 140])
    banner.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), verdict_bg),
        ("TOPPADDING",    (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ("LEFTPADDING",   (0, 0), (-1, -1), 14),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 14),
        ("FONTNAME",  (0, 0), (-1, -1), font_name),
        ("VALIGN",    (0, 0), (-1, -1), "MIDDLE"),
    ]))
    elements.append(banner)
    elements.append(Spacer(1, 6 * mm))

    # ── Basic info table ─────────────────────────────────────────
    image_hash = (report.get("image_hash") or "")
    hash_display = image_hash[:32] + ("..." if len(image_hash) > 32 else "")
    info_rows = [
        ["\u68c0\u6d4b\u7ed3\u8bba", report.get("conclusion", "")],
    ]
    if report.get("image_url"):
        url_val = (report["image_url"] or "")[:80]
        info_rows.append(["\u56fe\u7247\u6765\u6e90", url_val])
    info_rows += [
        ["\u5ba1\u6838\u5efa\u8bae", report.get("suggestion", "")],
        ["\u6a21\u578b\u7248\u672c", report.get("model_version", "")],
        ["\u56fe\u7247\u54c8\u5e0c", hash_display],
        ["\u68c0\u6d4b\u65f6\u95f4", report.get("created_at", "")],
    ]
    para_rows = [
        [Paragraph(str(k), styles["CN"]), Paragraph(str(v), styles["CN"])]
        for k, v in info_rows
    ]
    info_table = Table(para_rows, colWidths=col_w)
    info_table.setStyle(TableStyle([
        ("FONTNAME",  (0, 0), (-1, -1), font_name),
        ("FONTSIZE",  (0, 0), (-1, -1), 9.5),
        ("GRID",      (0, 0), (-1, -1), 0.4, colors.Color(0.82, 0.82, 0.82)),
        ("BACKGROUND",(0, 0), (0, -1), colors.Color(0.95, 0.95, 0.95)),
        ("VALIGN",    (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING",    (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LEFTPADDING",   (0, 0), (-1, -1), 10),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 8 * mm))

    # ── Probability breakdown ────────────────────────────────────
    probs = report.get("probs", [])
    if probs:
        elements.append(Paragraph("\u6982\u7387\u5206\u5e03", styles["CNSection"]))
        bar_fake = colors.Color(0.78, 0.18, 0.18)
        bar_real = colors.Color(0.10, 0.63, 0.30)
        prob_data = [["\u7c7b\u522b(\u4e2d)", "\u6982\u7387", "\u53ef\u89c6\u5316"]]
        for p in probs:
            score = p.get("score", 0)
            filled = int(score / 5)
            bar = "\u2588" * filled + "\u2591" * (20 - filled)
            prob_data.append([p.get("label_zh", ""), f'{score:.1f}%', bar])
        prob_table = Table(prob_data, colWidths=[90, 70, 300])
        prob_ts = [
            ("FONTNAME",  (0, 0), (-1, -1), font_name),
            ("FONTSIZE",  (0, 0), (-1, -1), 9),
            ("BACKGROUND",(0, 0), (-1,  0), colors.Color(0.90, 0.90, 0.90)),
            ("GRID",      (0, 0), (-1, -1), 0.3, colors.Color(0.85, 0.85, 0.85)),
            ("TOPPADDING",    (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("LEFTPADDING",   (0, 0), (-1, -1), 8),
        ]
        for idx, p in enumerate(probs, start=1):
            c = bar_fake if p.get("label", "") == "FAKE" else bar_real
            prob_ts.append(("TEXTCOLOR", (2, idx), (2, idx), c))
        prob_table.setStyle(TableStyle(prob_ts))
        elements.append(prob_table)
        elements.append(Spacer(1, 8 * mm))

    # ── Analysis & recommendation section ────────────────────────
    elements.append(Paragraph("\u68c0\u6d4b\u5206\u6790\u4e0e\u5efa\u8bae", styles["CNSection"]))
    explanation = report.get("explanation") or {}
    summary   = explanation.get("summary", "") or report.get("conclusion", "")
    clues     = explanation.get("clues") or []
    disclaimer = explanation.get("disclaimer", "")
    suggestion = report.get("suggestion", "")

    if summary:
        elements.append(Paragraph(f"\u7efc\u5408\u5224\u65ad\uff1a{summary}", styles["CN"]))
        elements.append(Spacer(1, 3 * mm))
    if clues:
        elements.append(Paragraph("\u68c0\u6d4b\u7279\u5f81\uff1a", styles["CN"]))
        for i, clue in enumerate(clues, 1):
            elements.append(Paragraph(f"\u3000{i}. {clue}", styles["CNSmall"]))
        elements.append(Spacer(1, 3 * mm))
    if suggestion:
        elements.append(Paragraph(f"\u5efa\u8bae\uff1a{suggestion}", styles["CN"]))
    if disclaimer:
        elements.append(Spacer(1, 3 * mm))
        elements.append(Paragraph(f"\u6ce8\u610f\uff1a{disclaimer}", styles["CNSmall"]))

    doc.build(
        elements,
        onFirstPage=_draw_page_decoration,
        onLaterPages=_draw_page_decoration,
    )
    return buf.getvalue()


def export_excel(report: dict) -> bytes:
    """Generate an Excel report with color-coded verdict and probability rows."""
    wb = Workbook()
    ws = wb.active
    ws.title = "\u68c0\u6d4b\u62a5\u544a"

    is_fake = (report.get("label", "") == "FAKE")
    _HEADER_FILL = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
    _FAKE_FILL   = PatternFill(start_color="FFDDDD", end_color="FFDDDD", fill_type="solid")
    _REAL_FILL   = PatternFill(start_color="DDFFDD", end_color="DDFFDD", fill_type="solid")
    _BOLD        = Font(bold=True)
    _FAKE_FONT   = Font(bold=True, color="CC0000")
    _REAL_FONT   = Font(bold=True, color="006600")

    verdict_fill = _FAKE_FILL if is_fake else _REAL_FILL
    verdict_font = _FAKE_FONT if is_fake else _REAL_FONT

    # Report title
    ws.append(["AIGI-Holmes \u68c0\u6d4b\u62a5\u544a"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])

    def _add_row(label, value, highlight=False):
        ws.append([label, str(value)])
        r = ws.max_row
        ws.cell(r, 1).font = _BOLD
        ws.cell(r, 1).fill = _HEADER_FILL
        if highlight:
            ws.cell(r, 2).fill = verdict_fill
            ws.cell(r, 2).font = verdict_font

    _add_row("\u68c0\u6d4b\u7ed3\u8bba", report.get("conclusion", ""))
    _add_row("\u6807\u7b7e",   "AI\u751f\u6210\u56fe\u50cf" if is_fake else "\u771f\u5b9e\u7167\u7247", highlight=True)
    _add_row("\u7f6e\u4fe1\u5ea6", f'{report.get("confidence", 0):.1f}%', highlight=True)
    _add_row("\u5ba1\u6838\u5efa\u8bae", report.get("suggestion", ""))
    _add_row("\u6a21\u578b\u7248\u672c", report.get("model_version", ""))
    _add_row("\u56fe\u7247\u54c8\u5e0c", report.get("image_hash", ""))
    _add_row("\u56fe\u7247 URL",  report.get("image_url", "") or "N/A")
    _add_row("\u68c0\u6d4b\u65f6\u95f4", report.get("created_at", ""))
    ws.append([])

    # Probability section
    ws.append(["\u6982\u7387\u5206\u5e03"])
    ws.cell(ws.max_row, 1).font = Font(bold=True, size=11)
    ws.append(["\u7c7b\u522b", "\u7c7b\u522b(\u4e2d\u6587)", "\u6982\u7387(%)"])
    for cell in ws[ws.max_row]:
        cell.font = _BOLD
        cell.fill = _HEADER_FILL

    for p in report.get("probs", []):
        score = p.get("score", 0)
        ws.append([p.get("label", ""), p.get("label_zh", ""), f'{score:.1f}'])
        r = ws.max_row
        row_fill = _FAKE_FILL if p.get("label", "") == "FAKE" else _REAL_FILL
        for col in range(1, 4):
            ws.cell(r, col).fill = row_fill

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 52
    ws.column_dimensions["C"].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
